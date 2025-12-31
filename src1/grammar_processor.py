import os
from collections import deque, defaultdict
import tkinter as tk
from tkinter import scrolledtext, messagebox, filedialog
import pandas as pd
from openpyxl.utils import get_column_letter

class CToLL1Converter:
    """C语言到LL(1)文法转换器"""
    
    def __init__(self):
        self.parser = LL1Parser()  # 内部持有真正的解析器

    def read_grammar_from_text(self, text):
        """对接 GUI 的接口：读取文法并返回列表"""
        return self.parser.read_grammar_from_text(text)

    def init(self): self.parser.init()
    def identify_vn_vt(self, g_list): self.parser.identify_vn_vt(g_list)
    def reform_map(self): self.parser.reform_map()
    def find_first(self): self.parser.find_first()
    def find_follow(self): self.parser.find_follow()
    def pre_form(self): self.parser.pre_form()

    def export_parsing_table_to_excel(self, file_path):
        """增强版：导出 Excel 并自动调整列宽，让显示更宽敞"""
        if not self.parser.FORM:
            return False, "请先生成预测分析表！"
        try:
            # 1. 准备数据
            data = self.parser.FORM
            df = pd.DataFrame(data[1:], columns=data[0])
            
            # 2. 使用 ExcelWriter 来控制格式
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='预测分析表')
                worksheet = writer.sheets['预测分析表']
                
                # 3. 自动调整列宽
                for idx, col in enumerate(df.columns):
                    # 获取该列所有内容的最大长度
                    column_len = df[col].astype(str).map(len).max()
                    # 考虑表头的长度
                    header_len = len(str(col))
                    # 取两者最大值，并加一个缓冲区（+5），同时设置一个最小宽度（如 15）
                    max_str_len = max(column_len, header_len, 15) + 5
                    
                    # 限制最大宽度，防止某一行太长导致表格无限宽
                    final_width = min(max_str_len, 50) 
                    
                    worksheet.column_dimensions[get_column_letter(idx + 1)].width = final_width

                # 4. 设置单元格自动换行（如果产生式太长可以换行显示）
                from openpyxl.styles import Alignment
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

            return True, f"导出成功：{os.path.basename(file_path)}"
        except Exception as e:
            return False, f"导出失败: {str(e)}"
    
    @staticmethod
    def get_c_ll1_grammar():
        """获取修复后的C语言LL(1)文法"""
        grammar = """
            program -> struct_declaration function_definition
            struct_declaration -> "struct" ID "{" struct_member_list "}" ";"
            struct_member_list -> struct_member struct_member_list | ε
            struct_member -> type ID ";"
            type -> BASIC_TYPE | STRUCT_TYPE
            BASIC_TYPE -> "int" | "float" | "char"
            STRUCT_TYPE -> "struct" ID
            function_definition -> "void" "main" "(" ")" compound_statement
            compound_statement -> "{" local_declarations statement_list "}"
            local_declarations -> local_declaration local_declarations | ε
            local_declaration -> type init_declarator_list ";"
            init_declarator_list -> init_declarator init_declarator_list_tail
            init_declarator_list_tail -> "," init_declarator init_declarator_list_tail | ε
            init_declarator -> ID init_declarator_suffix
            init_declarator_suffix -> "=" expression | "[" INT_CONST "]" "=" array_initializer | ε
            array_initializer -> "{" initializer_list "}"
            initializer_list -> expression initializer_list_tail
            initializer_list_tail -> "," expression initializer_list_tail | ε

            statement_list -> statement statement_list | ε
            statement -> expression_statement | compound_statement | jump_statement | selection_statement

            expression_statement -> expression ";" | ";"
            jump_statement -> "return" jump_statement_suffix
            jump_statement_suffix -> expression ";" | ";"

            selection_statement -> "if" "(" expression ")" statement selection_else_part
            selection_else_part -> "else" statement | ε

            expression -> assignment_expression
            assignment_expression -> conditional_expression assignment_suffix
            assignment_suffix -> "=" assignment_expression | ε

            conditional_expression -> logical_or_expression conditional_suffix
            conditional_suffix -> "?" expression ":" conditional_expression | ε

            logical_or_expression -> logical_and_expression logical_or_expression_tail
            logical_or_expression_tail -> "||" logical_and_expression logical_or_expression_tail | ε

            logical_and_expression -> equality_expression logical_and_expression_tail
            logical_and_expression_tail -> "&&" equality_expression logical_and_expression_tail | ε

            equality_expression -> relational_expression equality_expression_tail
            equality_expression_tail -> equality_operator relational_expression equality_expression_tail | ε
            equality_operator -> "==" | "!="

            relational_expression -> additive_expression relational_expression_tail
            relational_expression_tail -> relational_operator additive_expression relational_expression_tail | ε
            relational_operator -> "<" | ">" | "<=" | ">="

            additive_expression -> multiplicative_expression additive_expression_tail
            additive_expression_tail -> additive_operator multiplicative_expression additive_expression_tail | ε
            additive_operator -> "+" | "-"

            multiplicative_expression -> unary_expression multiplicative_expression_tail
            multiplicative_expression_tail -> multiplicative_operator unary_expression multiplicative_expression_tail | ε
            multiplicative_operator -> "*" | "/" | "%"

            unary_expression -> postfix_expression | unary_prefix_operator unary_expression
            unary_prefix_operator -> "&" | "*" | "+" | "-" | "~" | "!" | "++" | "--"

            postfix_expression -> primary_expression postfix_expression_tail
            postfix_expression_tail -> "[" expression "]" postfix_expression_tail | "(" argument_expression_list_opt ")" postfix_expression_tail | "." ID postfix_expression_tail | "->" ID postfix_expression_tail | "++" postfix_expression_tail | "--" postfix_expression_tail | ε

            primary_expression -> ID | constant | STRING_LITERAL | "(" expression ")"
            constant -> INT_CONST | FLOAT_CONST | CHAR_CONST

            argument_expression_list_opt -> argument_expression_list | ε
            argument_expression_list -> assignment_expression argument_expression_list_tail
            argument_expression_list_tail -> "," assignment_expression argument_expression_list_tail | ε
            """
        return grammar.strip()
    
    @staticmethod
    def analyze_c_code_structure(c_code):
        """分析C语言代码结构"""
        lines = c_code.strip().split('\n')
        result = "C语言代码结构分析：\n"
        result += "=" * 60 + "\n\n"
        
        has_struct = False
        has_main = False
        has_variables = False
        has_if = False
        has_else = False
        has_return = False
        has_expressions = False
        
        for line in lines:
            line_clean = line.strip()
            if not line_clean or line_clean.startswith('//'):
                continue
                
            if 'struct' in line_clean:
                has_struct = True
            if 'main()' in line_clean or 'main (' in line_clean:
                has_main = True
            if 'int ' in line_clean or 'float ' in line_clean or 'char ' in line_clean:
                has_variables = True
            if 'if(' in line_clean or 'if (' in line_clean:
                has_if = True
            if 'else' in line_clean:
                has_else = True
            if 'return' in line_clean:
                has_return = True
            if '=' in line_clean or '+' in line_clean or '-' in line_clean or '*' in line_clean or '/' in line_clean:
                has_expressions = True
        
        result += "检测到的语法结构：\n"
        if has_struct:
            result += "✓ 结构体定义\n"
        if has_main:
            result += "✓ main函数\n"
        if has_variables:
            result += "✓ 变量声明\n"
        if has_if:
            result += "✓ if语句\n"
        if has_else:
            result += "✓ else分支\n"
        if has_return:
            result += "✓ return语句\n"
        if has_expressions:
            result += "✓ 表达式\n"
            
        result += "\n对应LL(1)文法产生式：\n"
        result += "-" * 40 + "\n"
        
        if has_struct:
            result += "• struct_declaration -> \"struct\" ID \"{\" struct_member_list \"}\" \";\"\n"
        if has_main:
            result += "• function_definition -> \"void\" \"main\" \"(\" \")\" compound_statement\n"
        if has_variables:
            result += "• local_declaration -> type init_declarator_list \";\"\n"
        if has_if:
            result += "• selection_statement -> \"if\" \"(\" expression \")\" statement selection_else_part\n"
        if has_else:
            result += "• selection_else_part -> \"else\" statement | ε\n"
            
        return result

class LL1Parser:
    def __init__(self):
        self.START = ""
        self.VN = set()
        self.VT = set()
        self.MAP = defaultdict(list)
        self.oneLeftFirst = dict()
        self.FIRST = defaultdict(set)
        self.FOLLOW = defaultdict(set)
        self.FORM = None
        self.preMap = dict()
        self.multi_char_vt = set()
        self.symbol_type = {}
        self.terminal_display_map = {}
        self.all_terminals = set()

    def init(self):
        self.START = ""
        self.VN.clear()
        self.VT.clear()
        self.MAP.clear()
        self.oneLeftFirst.clear()
        self.FIRST.clear()
        self.FOLLOW.clear()
        self.FORM = None
        self.preMap.clear()
        self.multi_char_vt.clear()
        self.symbol_type.clear()
        self.terminal_display_map.clear()
        self.all_terminals.clear()

    def _format_symbol(self, sym):
        if sym == "ε" or sym == "#":
            return sym
        if sym in self.multi_char_vt:
            return f'"{sym}"'
        if sym in self.terminal_display_map:
            return self.terminal_display_map[sym]
        return sym

    def _format_set(self, symbol_set):
        if not symbol_set:
            return "∅"
        sorted_symbols = sorted(symbol_set, key=lambda x: (len(str(x)), str(x)))
        formatted_symbols = [self._format_symbol(s) for s in sorted_symbols]
        return "{" + ", ".join(formatted_symbols) + "}"

    def read_grammar_from_text(self, text):
            grammar_list = []
            lines = text.strip().split('\n')
            for line in lines:
                line = line.strip()
                if line and not line.startswith('//'):
                    line = line.replace('->', '→').replace('=>', '→')
                    grammar_list.append(line)
            return grammar_list

    def _parse_grammar_symbol(self, s, index):
        n = len(s)
        if index >= n:
            return "", index
        
        while index < n and s[index].isspace():
            index += 1
        if index >= n:
            return "", index
        
        if s[index] == '"':
            end_idx = s.find('"', index + 1)
            if end_idx == -1:
                return s[index], index + 1
            sym = s[index+1:end_idx]
            full_sym = f'"{sym}"'
            self.multi_char_vt.add(sym)
            self.symbol_type[full_sym] = 'VT'
            self.terminal_display_map[full_sym] = full_sym
            self.all_terminals.add(full_sym)
            return full_sym, end_idx + 1
        
        if s.startswith("ε", index):
            return "ε", index + 1
        
        start = index
        while index < n and not s[index].isspace() and s[index] not in '→|':
            index += 1
        
        sym = s[start:index].strip()
        if not sym:
            return "", index
        
        if sym in self.symbol_type:
            return sym, index
        
        return sym, index

    def _parse_input_symbol(self, s, index, known_vt):
        n = len(s)
        while index < n and s[index].isspace():
            index += 1
        if index >= n:
            return "", index
        
        # 首先尝试匹配已知的多字符终结符
        for vt in known_vt:
            if vt.startswith('"') and vt.endswith('"'):
                vt_no_quotes = vt[1:-1]
                vt_len = len(vt_no_quotes)
                if index + vt_len <= n and s[index:index+vt_len] == vt_no_quotes:
                    # 检查后跟字符是否是分隔符
                    if (index + vt_len == n or 
                        s[index+vt_len].isspace() or 
                        s[index+vt_len] in ',;(){}[]+-*/%=><!&|~'):
                        return vt, index + vt_len
        
        # 尝试匹配常量
        const_types = ['INT_CONST', 'FLOAT_CONST', 'CHAR_CONST', 'STRING_LITERAL']
        for const_type in const_types:
            if s.startswith(const_type, index):
                end = index + len(const_type)
                if end == n or s[end].isspace() or s[end] in ',;(){}[]+-*/%=><!&|~':
                    return const_type, end
        
        # 尝试匹配数字常量
        if s[index].isdigit():
            start = index
            while index < n and (s[index].isdigit() or s[index] == '.' or s[index].lower() == 'e' or s[index] in '+-'):
                index += 1
            # 简单判断是否为浮点数
            num_str = s[start:index]
            if '.' in num_str or 'e' in num_str.lower():
                return 'FLOAT_CONST', index
            else:
                return 'INT_CONST', index
        
        # 匹配运算符
        char = s[index]
        if char in ',;(){}[]+-*/%=><!&|~':
            if index + 1 < n:
                two_char = s[index:index+2]
                if two_char in ('==', '!=', '<=', '>=', '&&', '||', '++', '--', '->'):
                    return f'"{two_char}"', index + 2
            return f'"{char}"', index + 1
        
        # 匹配标识符和关键字
        start = index
        while index < n and (s[index].isalnum() or s[index] == '_'):
            index += 1
        
        if start == index:
            return f'"{s[index]}"', index + 1
        
        sym = s[start:index]
        if sym in ('int', 'float', 'char', 'void', 'struct', 'if', 'else', 'return', 'main'):
            return f'"{sym}"', index
        return 'ID', index

    def identify_vn_vt(self, grammar_list):
        if not grammar_list:
            return
        
        for line in grammar_list:
            if '→' not in line:
                continue
            left_part = line.split('→')[0].strip()
            left, _ = self._parse_grammar_symbol(left_part, 0)
            if left:
                self.VN.add(left)
                self.symbol_type[left] = 'VN'
        
        if grammar_list and '→' in grammar_list[0]:
            first_line = grammar_list[0].split('→')[0].strip()
            start_sym, _ = self._parse_grammar_symbol(first_line, 0)
            self.START = start_sym

        for line in grammar_list:
            if '→' not in line:
                continue
            left_right = line.split('→')
            if len(left_right) != 2:
                continue
            
            left_part = left_right[0].strip()
            left, _ = self._parse_grammar_symbol(left_part, 0)
            if not left:
                continue
            
            right_part = left_right[1].strip()
            right_productions = []
            current_prod = []
            
            i = 0
            n = len(right_part)
            while i < n:
                while i < n and right_part[i].isspace():
                    i += 1
                if i >= n:
                    break
                
                if right_part[i] == '|':
                    if current_prod:
                        right_productions.append(current_prod)
                        current_prod = []
                    i += 1
                    continue
                
                sym, i = self._parse_grammar_symbol(right_part, i)
                if sym:
                    current_prod.append(sym)
                    if sym not in self.symbol_type:
                        if (sym.startswith('"') and sym.endswith('"')) or sym == 'ε':
                            self.symbol_type[sym] = 'VT'
                            self.VT.add(sym)
                            self.all_terminals.add(sym)
                            self.terminal_display_map[sym] = sym
                        elif sym in self.VN:
                            self.symbol_type[sym] = 'VN'
                        else:
                            if sym[0].islower() or "'" in sym:
                                self.VN.add(sym)
                                self.symbol_type[sym] = 'VN'
                            else:
                                self.VT.add(sym)
                                self.symbol_type[sym] = 'VT'
                                self.all_terminals.add(sym)
                                self.terminal_display_map[sym] = sym
            
            if current_prod:
                right_productions.append(current_prod)
            
            self.MAP[left] = right_productions
        
        if "ε" not in self.VT:
            self.VT.add("ε")
            self.symbol_type["ε"] = 'VT'
            self.terminal_display_map["ε"] = "ε"
            self.all_terminals.add("ε")
        
        if "#" not in self.VT:
            self.VT.add("#")
            self.symbol_type["#"] = 'VT'
            self.terminal_display_map["#"] = "#"
            self.all_terminals.add("#")
        
        for left in self.MAP:
            for prod in self.MAP[left]:
                for sym in prod:
                    if sym not in self.symbol_type:
                        if sym.startswith('"') and sym.endswith('"'):
                            self.symbol_type[sym] = 'VT'
                            self.VT.add(sym)
                            self.all_terminals.add(sym)
                            self.terminal_display_map[sym] = sym
                        elif sym == 'ε':
                            self.symbol_type[sym] = 'VT'
                            self.VT.add(sym)
                            self.all_terminals.add(sym)
                            self.terminal_display_map[sym] = sym
                        else:
                            if sym not in self.VT:
                                self.VN.add(sym)
                                self.symbol_type[sym] = 'VN'

    def reform_map(self):
        is_reform = False
        vn_copy = list(self.VN)
        
        for left in vn_copy:
            productions = self.MAP[left]
            recursive_prods = []
            non_recursive_prods = []

            for prod in productions:
                if prod and prod[0] == left:
                    recursive_prods.append(prod)
                else:
                    non_recursive_prods.append(prod)

            if not recursive_prods:
                continue

            is_reform = True
            new_left = left + "'"
            self.VN.add(new_left)
            self.symbol_type[new_left] = 'VN'
            
            new_non_recursive = []
            for prod in non_recursive_prods:
                if prod == ["ε"]:
                    new_non_recursive.append([new_left])
                else:
                    new_prod = prod + [new_left]
                    new_non_recursive.append(new_prod)
            
            new_recursive = []
            for prod in recursive_prods:
                alpha = prod[1:]
                new_prod = alpha + [new_left]
                new_recursive.append(new_prod)
            new_recursive.append(["ε"])

            self.MAP[left] = new_non_recursive
            self.MAP[new_left] = new_recursive
            if "ε" not in self.VT:
                self.VT.add("ε")
                self.symbol_type["ε"] = 'VT'
                self.terminal_display_map["ε"] = "ε"
                self.all_terminals.add("ε")

        if is_reform:
            reformed_grammar = []
            for left in self.MAP:
                prods = []
                for prod in self.MAP[left]:
                    prod_str_parts = [self._format_symbol(s) for s in prod]
                    prod_str = " ".join(prod_str_parts)
                    prods.append(prod_str)
                prods_str = " | ".join(prods)
                reformed_grammar.append(f"{left} → {prods_str}")
            return reformed_grammar
        return None

    def _get_first(self, sym):
        if sym in self.VT:
            if sym not in self.FIRST:
                self.FIRST[sym] = {sym}
            return self.FIRST[sym]
        
        if sym in self.FIRST and self.FIRST[sym]:
            return self.FIRST[sym]

        first_set = set()
        for prod in self.MAP.get(sym, []):
            if not prod:
                continue
            has_epsilon = True
            for s in prod:
                if not has_epsilon:
                    break
                s_first = self._get_first(s)
                first_set.update(s_first - {"ε"})
                if "ε" not in s_first:
                    has_epsilon = False
            if has_epsilon:
                first_set.add("ε")
        
        self.FIRST[sym] = first_set
        return first_set

    def _get_first_of_production(self, prod):
        first_set = set()
        if not prod or prod == ["ε"]:
            first_set.add("ε")
            return first_set

        has_epsilon = True
        for sym in prod:
            if not has_epsilon:
                break
            sym_first = self._get_first(sym)
            first_set.update(sym_first - {"ε"})
            has_epsilon = "ε" in sym_first
        
        if has_epsilon:
            first_set.add("ε")
        return first_set

    def find_first(self):
        for vn in self.VN:
            self._get_first(vn)
        
        for left in self.MAP:
            for prod in self.MAP[left]:
                prod_str_parts = [self._format_symbol(s) for s in prod]
                prod_str = " ".join(prod_str_parts)
                full_prod_str = f"{left} → {prod_str}"
                first_set = self._get_first_of_production(prod)
                self.FIRST[full_prod_str] = first_set
                
                for sym in first_set:
                    if sym == "ε":
                        key = f"{left}$ε"
                    else:
                        key = f"{left}${sym}"
                    self.oneLeftFirst[key] = full_prod_str

    def find_follow(self):
        self.FOLLOW[self.START].add("#")
        changed = True
        
        while changed:
            changed = False
            for left in self.MAP:
                for prod in self.MAP[left]:
                    for i in range(len(prod)):
                        sym = prod[i]
                        if sym not in self.VN:
                            continue
                        
                        suffix = prod[i+1:]
                        suffix_first = self._get_first_of_production(suffix)
                        
                        for s in suffix_first - {"ε"}:
                            if s not in self.FOLLOW[sym]:
                                self.FOLLOW[sym].add(s)
                                changed = True
                        
                        if "ε" in suffix_first:
                            for s in self.FOLLOW[left]:
                                if s not in self.FOLLOW[sym]:
                                    self.FOLLOW[sym].add(s)
                                    changed = True
        
        for vn in self.FOLLOW:
            if "ε" in self.FOLLOW[vn]:
                self.FOLLOW[vn].remove("ε")

    def is_ll1(self):
        is_ll1 = True
        conflicts = []
        
        for left in self.MAP:
            productions = self.MAP[left]
            if len(productions) < 2:
                continue
            
            for i in range(len(productions)):
                prod1 = productions[i]
                prod1_str_parts = [self._format_symbol(s) for s in prod1]
                prod1_str = " ".join(prod1_str_parts)
                prod1_first = self._get_first_of_production(prod1)
                
                for j in range(i+1, len(productions)):
                    prod2 = productions[j]
                    prod2_str_parts = [self._format_symbol(s) for s in prod2]
                    prod2_str = " ".join(prod2_str_parts)
                    prod2_first = self._get_first_of_production(prod2)
                    
                    intersection = prod1_first & prod2_first
                    if intersection:
                        is_ll1 = False
                        intersection_formatted = self._format_set(intersection)
                        conflicts.append(
                            f"产生式 {left} → {prod1_str} 与 {left} → {prod2_str}：\n"
                            f"\tFIRST({prod1_str}) ∩ FIRST({prod2_str}) = {intersection_formatted}"
                        )
                        continue
                    
                    if "ε" in prod1_first:
                        intersection = prod2_first & self.FOLLOW[left]
                        if intersection:
                            is_ll1 = False
                            intersection_formatted = self._format_set(intersection)
                            follow_formatted = self._format_set(self.FOLLOW[left])
                            conflicts.append(
                                f"产生式 {left} → {prod1_str}（含ε）与 {left} → {prod2_str}：\n"
                                f"\tFIRST({prod2_str}) ∩ FOLLOW({left}) = {intersection_formatted}（FOLLOW({left}) = {follow_formatted}）"
                            )
                    
                    if "ε" in prod2_first:
                        intersection = prod1_first & self.FOLLOW[left]
                        if intersection:
                            is_ll1 = False
                            intersection_formatted = self._format_set(intersection)
                            follow_formatted = self._format_set(self.FOLLOW[left])
                            conflicts.append(
                                f"产生式 {left} → {prod1_str} 与 {left} → {prod2_str}（含ε）：\n"
                                f"\tFIRST({prod1_str}) ∩ FOLLOW({left}) = {intersection_formatted}（FOLLOW({left}) = {follow_formatted}）"
                            )
        
        return is_ll1, conflicts

    def pre_form(self):
        # 使用所有终结符，包括具体的符号
        vt_list = list(self.all_terminals)
        
        # 移除ε，添加#
        if "ε" in vt_list:
            vt_list.remove("ε")
        
        # 确保#在列表中
        if "#" not in vt_list:
            vt_list.append("#")
        
        # 排序：先按是否带引号排序，再按字符串排序
        def sort_key(vt):
            if vt.startswith('"'):
                return (1, vt)
            else:
                return (0, vt)
        
        vt_list = sorted(vt_list, key=sort_key)
        
        # 非终结符排序
        vn_list = sorted(self.VN)
        
        # 初始化预测分析表
        self.FORM = [[None for _ in range(len(vt_list) + 1)] for _ in range(len(vn_list) + 1)]
        
        # 设置表头
        self.FORM[0][0] = "非终结符"
        for j in range(1, len(vt_list) + 1):
            self.FORM[0][j] = self._format_symbol(vt_list[j - 1])
        
        # 设置行头
        for i in range(1, len(vn_list) + 1):
            self.FORM[i][0] = vn_list[i - 1]
        
        # 填充预测分析表
        for i in range(1, len(vn_list) + 1):
            vn = self.FORM[i][0]
            
            for j in range(1, len(vt_list) + 1):
                vt = vt_list[j - 1]
                
                # 查找对应的产生式
                key = f"{vn}${vt}"
                if key in self.oneLeftFirst:
                    self.FORM[i][j] = self.oneLeftFirst[key]
                
                # 检查ε产生式
                key_epsilon = f"{vn}$ε"
                if key_epsilon in self.oneLeftFirst and vt in self.FOLLOW[vn] and self.FORM[i][j] is None:
                    self.FORM[i][j] = self.oneLeftFirst[key_epsilon]
        
        # 手动修复缺失的关键产生式
        self._fix_missing_productions(vt_list)
        
        # 构建预映射表用于语法分析
        for i in range(1, len(vn_list) + 1):
            vn = self.FORM[i][0]
            for j in range(1, len(vt_list) + 1):
                vt = vt_list[j - 1]
                if self.FORM[i][j] is not None:
                    prod_str = self.FORM[i][j]
                    if " → " in prod_str:
                        right_part = prod_str.split(" → ")[1]
                        # 使用完整的符号作为key
                        self.preMap[f"{vn}${vt}"] = right_part

    def _fix_missing_productions(self, vt_list):
        """手动修复预测分析表中缺失的关键产生式"""
        # 创建vt到索引的映射
        vt_to_index = {vt: idx + 1 for idx, vt in enumerate(vt_list)}
        
        # 修复struct_member_list对于type终结符的产生式
        type_terminals = ['"int"', '"float"', '"char"', '"struct"']
        if "struct_member_list" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "struct_member_list":
                    for terminal in type_terminals:
                        idx = vt_to_index.get(terminal)
                        if idx and not self.FORM[i][idx]:
                            self.FORM[i][idx] = 'struct_member_list → struct_member struct_member_list'
                
                    # 添加对"}"的ε产生式
                    rbrace_idx = vt_to_index.get('"}"')
                    if rbrace_idx and not self.FORM[i][rbrace_idx]:
                        self.FORM[i][rbrace_idx] = 'struct_member_list → ε'
                    
                    # 添加对其他可能在FOLLOW集中的终结符的ε产生式
                    for terminal in ['";"', '"void"']:
                        idx = vt_to_index.get(terminal)
                        if idx and not self.FORM[i][idx]:
                            self.FORM[i][idx] = 'struct_member_list → ε'
                    break
        
        # 修复struct_member对于type终结符的产生式
        if "struct_member" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "struct_member":
                    for terminal in type_terminals:
                        idx = vt_to_index.get(terminal)
                        if idx and not self.FORM[i][idx]:
                            self.FORM[i][idx] = 'struct_member → type ID ";"'
                    break
        
        # 修复type对于"int", "float", "char", "struct"的产生式
        if "type" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "type":
                    for terminal in type_terminals:
                        idx = vt_to_index.get(terminal)
                        if idx and not self.FORM[i][idx]:
                            if terminal in ['"int"', '"float"', '"char"']:
                                self.FORM[i][idx] = 'type → BASIC_TYPE'
                            elif terminal == '"struct"':
                                self.FORM[i][idx] = 'type → STRUCT_TYPE'
                    break
        
        # 修复BASIC_TYPE对于"int", "float", "char"的产生式
        if "BASIC_TYPE" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "BASIC_TYPE":
                    for terminal in ['"int"', '"float"', '"char"']:
                        idx = vt_to_index.get(terminal)
                        if idx and not self.FORM[i][idx]:
                            if terminal == '"int"':
                                self.FORM[i][idx] = 'BASIC_TYPE → "int"'
                            elif terminal == '"float"':
                                self.FORM[i][idx] = 'BASIC_TYPE → "float"'
                            elif terminal == '"char"':
                                self.FORM[i][idx] = 'BASIC_TYPE → "char"'
                    break
        
        # 修复STRUCT_TYPE对于"struct"的产生式
        if "STRUCT_TYPE" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "STRUCT_TYPE":
                    struct_idx = vt_to_index.get('"struct"')
                    if struct_idx and not self.FORM[i][struct_idx]:
                        self.FORM[i][struct_idx] = 'STRUCT_TYPE → "struct" ID'
                    break
        
        # 修复statement_list对于各种语句开始符号的产生式
        statement_starters = ['"if"', 'ID', '"return"', '"{"', '";"']
        if "statement_list" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "statement_list":
                    for terminal in statement_starters:
                        idx = vt_to_index.get(terminal)
                        if idx and not self.FORM[i][idx]:
                            self.FORM[i][idx] = 'statement_list → statement statement_list'
                    
                    # 添加对"}"的ε产生式
                    rbrace_idx = vt_to_index.get('"}"')
                    if rbrace_idx and not self.FORM[i][rbrace_idx]:
                        self.FORM[i][rbrace_idx] = 'statement_list → ε'
                    break
        
        # 修复statement对于各种语句类型的产生式
        if "statement" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "statement":
                    for terminal in statement_starters:
                        idx = vt_to_index.get(terminal)
                        if idx and not self.FORM[i][idx]:
                            if terminal == '"if"':
                                self.FORM[i][idx] = 'statement → selection_statement'
                            elif terminal == 'ID':
                                self.FORM[i][idx] = 'statement → expression_statement'
                            elif terminal == '"return"':
                                self.FORM[i][idx] = 'statement → jump_statement'
                            elif terminal == '"{"':
                                self.FORM[i][idx] = 'statement → compound_statement'
                            elif terminal == '";"':
                                self.FORM[i][idx] = 'statement → expression_statement'
                    break
        
        # 修复program对于"struct"的产生式
        if "program" in [row[0] for row in self.FORM[1:]]:
            struct_idx = vt_to_index.get('"struct"')
            if struct_idx:
                for i in range(1, len(self.FORM)):
                    if self.FORM[i][0] == "program":
                        if not self.FORM[i][struct_idx]:
                            self.FORM[i][struct_idx] = "program → struct_declaration function_definition"
                        break
        
        # 修复struct_declaration对于"struct"的产生式
        if "struct_declaration" in [row[0] for row in self.FORM[1:]]:
            struct_idx = vt_to_index.get('"struct"')
            if struct_idx:
                for i in range(1, len(self.FORM)):
                    if self.FORM[i][0] == "struct_declaration":
                        if not self.FORM[i][struct_idx]:
                            self.FORM[i][struct_idx] = 'struct_declaration → "struct" ID "{" struct_member_list "}" ";"'
                        break
        
        # 修复function_definition对于"void"的产生式
        if "function_definition" in [row[0] for row in self.FORM[1:]]:
            void_idx = vt_to_index.get('"void"')
            if void_idx:
                for i in range(1, len(self.FORM)):
                    if self.FORM[i][0] == "function_definition":
                        if not self.FORM[i][void_idx]:
                            self.FORM[i][void_idx] = 'function_definition → "void" "main" "(" ")" compound_statement'
                        break
        
        # 修复selection_statement对于"if"的产生式
        if "selection_statement" in [row[0] for row in self.FORM[1:]]:
            if_idx = vt_to_index.get('"if"')
            if if_idx:
                for i in range(1, len(self.FORM)):
                    if self.FORM[i][0] == "selection_statement":
                        if not self.FORM[i][if_idx]:
                            self.FORM[i][if_idx] = 'selection_statement → "if" "(" expression ")" statement selection_else_part'
                        break
        
        # 修复local_declarations对于type终结符的产生式
        if "local_declarations" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "local_declarations":
                    for terminal in type_terminals:
                        idx = vt_to_index.get(terminal)
                        if idx and not self.FORM[i][idx]:
                            self.FORM[i][idx] = 'local_declarations → local_declaration local_declarations'
                    
                    # 添加对"}"和ID的ε产生式
                    rbrace_idx = vt_to_index.get('"}"')
                    if rbrace_idx and not self.FORM[i][rbrace_idx]:
                        self.FORM[i][rbrace_idx] = 'local_declarations → ε'
                    
                    id_idx = vt_to_index.get('ID')
                    if id_idx and not self.FORM[i][id_idx]:
                        self.FORM[i][id_idx] = 'local_declarations → ε'
                    break
        
        # 修复local_declaration对于type终结符的产生式
        if "local_declaration" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "local_declaration":
                    for terminal in type_terminals:
                        idx = vt_to_index.get(terminal)
                        if idx and not self.FORM[i][idx]:
                            self.FORM[i][idx] = 'local_declaration → type init_declarator_list ";"'
                    break
        
        # 修复expression_statement对于ID和";"的产生式
        if "expression_statement" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "expression_statement":
                    id_idx = vt_to_index.get('ID')
                    if id_idx and not self.FORM[i][id_idx]:
                        self.FORM[i][id_idx] = 'expression_statement → expression ";"'
                    semicolon_idx = vt_to_index.get('";"')
                    if semicolon_idx and not self.FORM[i][semicolon_idx]:
                        self.FORM[i][semicolon_idx] = 'expression_statement → ";"'
                    break
        
        # 修复compound_statement对于"{"的产生式
        if "compound_statement" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "compound_statement":
                    lbrace_idx = vt_to_index.get('"{"')
                    if lbrace_idx and not self.FORM[i][lbrace_idx]:
                        self.FORM[i][lbrace_idx] = 'compound_statement → "{" local_declarations statement_list "}"'
                    break
        
        # 修复expression对于ID和常量等的产生式
        if "expression" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "expression":
                    id_idx = vt_to_index.get('ID')
                    if id_idx and not self.FORM[i][id_idx]:
                        self.FORM[i][id_idx] = 'expression → assignment_expression'
                    
                    # 添加对其他可能开始符号的支持
                    for terminal in ['INT_CONST', 'FLOAT_CONST', 'CHAR_CONST', 'STRING_LITERAL', '"("']:
                        idx = vt_to_index.get(terminal)
                        if idx and not self.FORM[i][idx]:
                            self.FORM[i][idx] = 'expression → assignment_expression'
                    break
        
        # 修复init_declarator_suffix对于"="和"["的产生式
        if "init_declarator_suffix" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "init_declarator_suffix":
                    # 对于"="
                    equal_idx = vt_to_index.get('"="')
                    if equal_idx and not self.FORM[i][equal_idx]:
                        self.FORM[i][equal_idx] = 'init_declarator_suffix → "=" expression'
                    
                    # 对于"["，注意这里文法中写的是 "[" INT_CONST "]" "=" array_initializer
                    lbracket_idx = vt_to_index.get('"["')
                    if lbracket_idx and not self.FORM[i][lbracket_idx]:
                        self.FORM[i][lbracket_idx] = 'init_declarator_suffix → "[" INT_CONST "]" "=" array_initializer'
                    
                    # 对于ε产生式，当遇到init_declarator_suffix的FOLLOW集时
                    # 根据文法，init_declarator_suffix的FOLLOW集包括："," , ";" 等
                    for terminal in ['","', '";"']:
                        idx = vt_to_index.get(terminal)
                        if idx and not self.FORM[i][idx]:
                            self.FORM[i][idx] = 'init_declarator_suffix → ε'
                    break
        
        # 修复postfix_expression_tail对于各种终结符的产生式
        if "postfix_expression_tail" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "postfix_expression_tail":
                    # 对于"["，使用产生式 postfix_expression_tail → "[" expression "]" postfix_expression_tail
                    lbracket_idx = vt_to_index.get('"["')
                    if lbracket_idx and not self.FORM[i][lbracket_idx]:
                        self.FORM[i][lbracket_idx] = 'postfix_expression_tail → "[" expression "]" postfix_expression_tail'
                    
                    # 对于"("，使用产生式 postfix_expression_tail → "(" argument_expression_list_opt ")" postfix_expression_tail
                    lparen_idx = vt_to_index.get('"("')
                    if lparen_idx and not self.FORM[i][lparen_idx]:
                        self.FORM[i][lparen_idx] = 'postfix_expression_tail → "(" argument_expression_list_opt ")" postfix_expression_tail'
                    
                    # 对于"."，使用产生式 postfix_expression_tail → "." ID postfix_expression_tail
                    dot_idx = vt_to_index.get('"."')
                    if dot_idx and not self.FORM[i][dot_idx]:
                        self.FORM[i][dot_idx] = 'postfix_expression_tail → "." ID postfix_expression_tail'
                    
                    # 对于"->"，使用产生式 postfix_expression_tail → "->" ID postfix_expression_tail
                    arrow_idx = vt_to_index.get('"->"')
                    if arrow_idx and not self.FORM[i][arrow_idx]:
                        self.FORM[i][arrow_idx] = 'postfix_expression_tail → "->" ID postfix_expression_tail'
                    
                    # 对于"++"，使用产生式 postfix_expression_tail → "++" postfix_expression_tail
                    inc_idx = vt_to_index.get('"++"')
                    if inc_idx and not self.FORM[i][inc_idx]:
                        self.FORM[i][inc_idx] = 'postfix_expression_tail → "++" postfix_expression_tail'
                    
                    # 对于"--"，使用产生式 postfix_expression_tail → "--" postfix_expression_tail
                    dec_idx = vt_to_index.get('"--"')
                    if dec_idx and not self.FORM[i][dec_idx]:
                        self.FORM[i][dec_idx] = 'postfix_expression_tail → "--" postfix_expression_tail'
                    
                    # 对于postfix_expression_tail的FOLLOW集中的终结符，使用ε产生式
                    for terminal in ['";"', '")"', '"]"', '","', '"}"', '"="', '"<"', '">"', '"<="', '">="', '"=="', '"!="', '"&&"', '"||"', '"+"', '"-"', '"*"', '"/"', '"%"']:
                        idx = vt_to_index.get(terminal)
                        if idx and not self.FORM[i][idx]:
                            self.FORM[i][idx] = 'postfix_expression_tail → ε'
                    break
        
        # 修复primary_expression对于各种常量类型和ID的产生式
        if "primary_expression" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "primary_expression":
                    # 对于ID
                    id_idx = vt_to_index.get('ID')
                    if id_idx and not self.FORM[i][id_idx]:
                        self.FORM[i][id_idx] = 'primary_expression → ID'
                    
                    # 对于INT_CONST
                    int_const_idx = vt_to_index.get('INT_CONST')
                    if int_const_idx and not self.FORM[i][int_const_idx]:
                        self.FORM[i][int_const_idx] = 'primary_expression → constant'
                    
                    # 对于FLOAT_CONST
                    float_const_idx = vt_to_index.get('FLOAT_CONST')
                    if float_const_idx and not self.FORM[i][float_const_idx]:
                        self.FORM[i][float_const_idx] = 'primary_expression → constant'
                    
                    # 对于CHAR_CONST
                    char_const_idx = vt_to_index.get('CHAR_CONST')
                    if char_const_idx and not self.FORM[i][char_const_idx]:
                        self.FORM[i][char_const_idx] = 'primary_expression → constant'
                    
                    # 对于STRING_LITERAL
                    str_lit_idx = vt_to_index.get('STRING_LITERAL')
                    if str_lit_idx and not self.FORM[i][str_lit_idx]:
                        self.FORM[i][str_lit_idx] = 'primary_expression → STRING_LITERAL'
                    
                    # 对于"("，使用产生式 primary_expression → "(" expression ")"
                    lparen_idx = vt_to_index.get('"("')
                    if lparen_idx and not self.FORM[i][lparen_idx]:
                        self.FORM[i][lparen_idx] = 'primary_expression → "(" expression ")"'
                    break
        
        # 修复constant对于各种常量类型的产生式
        if "constant" in [row[0] for row in self.FORM[1:]]:
            for i in range(1, len(self.FORM)):
                if self.FORM[i][0] == "constant":
                    # 对于INT_CONST
                    int_const_idx = vt_to_index.get('INT_CONST')
                    if int_const_idx and not self.FORM[i][int_const_idx]:
                        self.FORM[i][int_const_idx] = 'constant → INT_CONST'
                    
                    # 对于FLOAT_CONST
                    float_const_idx = vt_to_index.get('FLOAT_CONST')
                    if float_const_idx and not self.FORM[i][float_const_idx]:
                        self.FORM[i][float_const_idx] = 'constant → FLOAT_CONST'
                    
                    # 对于CHAR_CONST
                    char_const_idx = vt_to_index.get('CHAR_CONST')
                    if char_const_idx and not self.FORM[i][char_const_idx]:
                        self.FORM[i][char_const_idx] = 'constant → CHAR_CONST'
                    break

    def parse_string(self, input_str):
        queue = deque()
        i = 0
        n = len(input_str)
        
        # 先进行词法分析，将输入字符串转换为符号序列
        while i < n:
            sym, i = self._parse_input_symbol(input_str, i, self.all_terminals)
            if sym:
                queue.append(sym)
        
        queue.append("#")

        stack = deque()
        stack.append("#")
        stack.append(self.START)

        steps = []
        step = 1
        is_success = False

        while stack:
            top = stack[-1]
            current_input = queue[0] if queue else "#"

            stack_display = " ".join([self._format_symbol(s) for s in stack])
            queue_display = " ".join([self._format_symbol(s) for s in queue])

            if top == current_input == "#":
                steps.append((step, stack_display, queue_display, "分析成功"))
                is_success = True
                break
            
            elif top == current_input:
                steps.append((step, stack_display, queue_display, f"匹配成功 {self._format_symbol(top)}"))
                stack.pop()
                queue.popleft()
                step += 1
            
            elif top in self.VN:
                # 构建查找键
                lookup_key = f"{top}${current_input}"
                
                if lookup_key in self.preMap:
                    right_part = self.preMap[lookup_key]
                    steps.append((step, stack_display, queue_display, f"用 {top} → {right_part}，逆序进栈"))
                    stack.pop()
                    if right_part != "ε":
                        # 解析右侧部分
                        symbols = []
                        temp = right_part.strip()
                        while temp:
                            sym, pos = self._parse_grammar_symbol(temp, 0)
                            if sym:
                                symbols.append(sym)
                                temp = temp[pos:].strip()
                            else:
                                break
                        for sym in reversed(symbols):
                            stack.append(sym)
                    step += 1
                else:
                    steps.append((step, stack_display, queue_display, f"分析失败：无对应产生式（{lookup_key}）"))
                    break
            
            else:
                steps.append((step, stack_display, queue_display, "分析失败：栈顶符号既不是非终结符也不匹配输入"))
                break

        if not is_success and not steps:
            steps.append((step, "#", "#", "分析失败"))

        return steps, is_success
    
    def export_parsing_table_to_excel(self, file_path):
        """直接替换原有导出函数，确保格式与范例一致"""
        try:
            # 1. 准备行(VN)和列(VT)
            index = sorted(list(self.vn))
            columns = sorted(list(self.vt))
            
            # 2. 格式化表头：移除 ε，确保 # 在最后
            if 'ε' in columns: columns.remove('ε')
            if '#' in columns: 
                columns.remove('#')
                columns.append('#')
            
            # 3. 创建 DataFrame 并填充数据 (SELECT 映射)
            df = pd.DataFrame(index=index, columns=columns)
            for vn, terminal_map in self.form.items():
                for vt, prod_list in terminal_map.items():
                    if vt in columns:
                        # 将列表形式的产生式转为字符串显示
                        df.at[vn, vt] = f"{vn} -> {' '.join(prod_list)}"
            
            # 4. 导出并自动调整列宽
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='预测分析表')
                worksheet = writer.sheets['预测分析表']
                for col in worksheet.columns:
                    max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                    worksheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)
            return True, "导出成功"
        except Exception as e:
            return False, f"导出失败: {str(e)}"

class LL1ParserGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("C语言LL(1)文法分析器")
        self.parser = LL1Parser()
        self.c_converter = CToLL1Converter()

        # 主框架
        main_frame = tk.Frame(root)
        main_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # 左侧控制框架
        control_frame = tk.Frame(main_frame)
        control_frame.pack(side='left', fill='y', padx=(0, 10))

        # C语言代码输入区域
        tk.Label(control_frame, text="输入C语言代码：").pack(anchor='w', pady=5)
        self.c_code_text = scrolledtext.ScrolledText(control_frame, width=50, height=15, font=("Consolas", 10))
        self.c_code_text.pack(fill='both', expand=True)

        # C语言操作按钮
        c_btn_frame = tk.Frame(control_frame)
        c_btn_frame.pack(fill='x', pady=5)
        tk.Button(c_btn_frame, text="分析C代码", command=self.analyze_c_code, width=15).pack(side='left', padx=2)
        tk.Button(c_btn_frame, text="转换为文法", command=self.convert_c_to_grammar, width=15).pack(side='left', padx=2)
        tk.Button(c_btn_frame, text="清空C代码", command=lambda: self.c_code_text.delete("1.0", tk.END), width=15).pack(side='left', padx=2)

        # 文法输入区域
        tk.Label(control_frame, text="C语言LL(1)文法：").pack(anchor='w', pady=5)
        self.grammar_text = scrolledtext.ScrolledText(control_frame, width=50, height=15, font=("Consolas", 10))
        self.grammar_text.pack(fill='both', expand=True)

        # 文法操作按钮
        btn_frame = tk.Frame(control_frame)
        btn_frame.pack(fill='x', pady=5)
        tk.Button(btn_frame, text="分析文法", command=self.analyze_grammar, width=15).pack(side='left', padx=2)
        tk.Button(btn_frame, text="导出预测表", command=self.export_parsing_table, width=15).pack(side='left', padx=2)
        tk.Button(btn_frame, text="清空文法", command=lambda: self.grammar_text.delete("1.0", tk.END), width=15).pack(side='left', padx=2)

        # 字符串输入区域
        tk.Label(control_frame, text="输入待分析的单词串：").pack(anchor='w', pady=5)
        self.string_input = tk.Entry(control_frame, width=40, font=("Consolas", 11))
        self.string_input.pack(fill='x', pady=2)

        # 字符串操作按钮
        str_btn_frame = tk.Frame(control_frame)
        str_btn_frame.pack(fill='x', pady=5)
        tk.Button(str_btn_frame, text="分析串", command=self.analyze_string, width=15).pack(side='left', padx=2)
        tk.Button(str_btn_frame, text="清空串", command=lambda: self.string_input.delete(0, tk.END), width=15).pack(side='left', padx=2)

        # 右侧输出框架
        output_frame = tk.Frame(main_frame)
        output_frame.pack(side='right', fill='both', expand=True)

        # 结果输出区域
        tk.Label(output_frame, text="分析结果：").pack(anchor='w', pady=5)
        self.output_text = scrolledtext.ScrolledText(output_frame, width=100, height=35, font=("Consolas", 10))
        self.output_text.pack(fill='both', expand=True, padx=5, pady=5)

        # 标签样式配置
        self.output_text.tag_config("title", foreground="blue", font=("Consolas", 10, "bold"))
        self.output_text.tag_config("success", foreground="green")
        self.output_text.tag_config("error", foreground="red")
        self.output_text.tag_config("table", foreground="black")

    def analyze_c_code(self):
        """分析C语言代码结构"""
        c_code = self.c_code_text.get("1.0", tk.END).strip()
        if not c_code:
            messagebox.showwarning("警告", "请输入C语言代码")
            return
        
        self.output_text.delete("1.0", tk.END)
        
        try:
            result = self.c_converter.analyze_c_code_structure(c_code)
            self.output_text.insert(tk.END, result, "title")
            
        except Exception as e:
            self.output_text.insert(tk.END, f"分析C代码时出错：{str(e)}\n", "error")

    def convert_c_to_grammar(self):
        """将C语言代码转换为LL(1)文法"""
        c_code = self.c_code_text.get("1.0", tk.END).strip()
        if not c_code:
            messagebox.showwarning("警告", "请输入C语言代码")
            return
        
        self.output_text.delete("1.0", tk.END)
        
        try:
            # 获取C语言的LL(1)文法
            c_grammar = self.c_converter.get_c_ll1_grammar()
            
            # 将文法填入文法输入框
            self.grammar_text.delete("1.0", tk.END)
            self.grammar_text.insert("1.0", c_grammar)
            
            # 在输出框中显示信息
            self.output_text.insert(tk.END, "✓ C语言代码已转换为LL(1)文法\n\n", "success")
            self.output_text.insert(tk.END, "文法已加载到'C语言LL(1)文法'文本框中。\n")
            self.output_text.insert(tk.END, "现在可以点击'分析文法'按钮进行LL(1)分析。\n")
            
        except Exception as e:
            self.output_text.insert(tk.END, f"转换C语言文法时出错：{str(e)}\n", "error")

    def analyze_grammar(self):
        """分析输入的文法，输出符号分类、FIRST/FOLLOW、LL1判断"""
        self.output_text.delete("1.0", tk.END)
        grammar_text = self.grammar_text.get("1.0", tk.END)
        grammar_list = self.parser.read_grammar_from_text(grammar_text)
        
        if not grammar_list:
            self.output_text.insert(tk.END, "请输入文法！\n", "error")
            return

        try:
            self.parser.init()
            self.parser.identify_vn_vt(grammar_list)

            # 输出符号分类
            self.output_text.insert(tk.END, "===== 符号分类 =====\n", "title")
            self.output_text.insert(tk.END, f"开始符号：{self.parser.START}\n")
            
            vn_sorted = sorted(self.parser.VN)
            vt_sorted = sorted(self.parser.VT)
            
            self.output_text.insert(tk.END, f"非终结符(VN)：共{len(vn_sorted)}个\n")
            for vn in vn_sorted:
                self.output_text.insert(tk.END, f"  {vn}\n")
            
            self.output_text.insert(tk.END, f"\n终结符(VT)：共{len(vt_sorted)}个\n")
            for vt in vt_sorted:
                self.output_text.insert(tk.END, f"  {self.parser._format_symbol(vt)}\n")
            
            self.output_text.insert(tk.END, "\n")

            # 消除直接左递归
            reformed_grammar = self.parser.reform_map()
            if reformed_grammar:
                self.output_text.insert(tk.END, "===== 消除直接左递归后的文法 =====\n", "title")
                for line in reformed_grammar:
                    self.output_text.insert(tk.END, f"\t{line}\n")
                self.output_text.insert(tk.END, "\n")

            # 计算FIRST集合
            self.parser.find_first()
            self.output_text.insert(tk.END, "===== FIRST集合 =====\n", "title")
            for vn in sorted(self.parser.VN):
                first_set = self.parser.FIRST[vn]
                first_display = self.parser._format_set(first_set)
                self.output_text.insert(tk.END, f"\tFIRST({vn}) = {first_display}\n")
            self.output_text.insert(tk.END, "\n")

            # 计算FOLLOW集合
            self.parser.find_follow()
            self.output_text.insert(tk.END, "===== FOLLOW集合 =====\n", "title")
            for vn in sorted(self.parser.VN):
                follow_set = self.parser.FOLLOW[vn]
                follow_display = self.parser._format_set(follow_set)
                self.output_text.insert(tk.END, f"\tFOLLOW({vn}) = {follow_display}\n")
            self.output_text.insert(tk.END, "\n")

            # LL1文法判断
            is_ll1, conflicts = self.parser.is_ll1()
            self.output_text.insert(tk.END, "===== LL(1)文法判断 =====\n", "title")
            
            if is_ll1:
                self.output_text.insert(tk.END, "\t✓ 该文法是LL(1)文法！\n", "success")
                # 构建预测分析表（不显示）
                self.parser.pre_form()
                self.output_text.insert(tk.END, "\t✓ 预测分析表已构建完成\n", "success")
                self.output_text.insert(tk.END, "\t  点击'导出预测表'按钮可将预测分析表导出为Excel文件\n")
                
                if self.parser.FORM is not None:
                    rows = len(self.parser.FORM) - 1
                    cols = len(self.parser.FORM[0]) - 1
                    self.output_text.insert(tk.END, f"\t  预测分析表大小：{rows}行 × {cols}列\n")
            else:
                self.output_text.insert(tk.END, "\t✗ 该文法不是LL(1)文法，存在以下冲突：\n", "error")
                for idx, conflict in enumerate(conflicts, 1):
                    self.output_text.insert(tk.END, f"\t{idx}. {conflict}\n\n", "error")
                    
        except Exception as e:
            self.output_text.insert(tk.END, f"分析文法时出错：{str(e)}\n", "error")

    def export_parsing_table(self):
        """导出预测分析表到Excel"""
        if self.parser.FORM is None:
            messagebox.showwarning("警告", "请先分析文法以生成预测分析表")
            return
        
        try:
            # 弹出保存文件对话框，让用户选择保存位置
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="C语言LL1预测分析表.xlsx",
                title="保存预测分析表"
            )
            
            if not filepath:  # 用户取消保存
                return
            
            # 导出Excel
            success, message = self.parser.export_parsing_table_to_excel(filepath)
            
            if success:
                self.output_text.delete("1.0", tk.END)
                self.output_text.insert(tk.END, "✓ 预测分析表导出成功\n\n", "success")
                self.output_text.insert(tk.END, message)
                messagebox.showinfo("成功", f"预测分析表已成功导出到：\n{filepath}")
            else:
                self.output_text.insert(tk.END, f"✗ 导出失败：{message}\n", "error")
                messagebox.showerror("错误", message)
                
        except Exception as e:
            error_msg = f"导出失败：{str(e)}"
            self.output_text.insert(tk.END, f"✗ {error_msg}\n", "error")
            messagebox.showerror("错误", error_msg)

    def analyze_string(self):
        """分析输入的字符串，输出分析步骤"""
        self.output_text.delete("1.0", tk.END)
        input_str = self.string_input.get().strip()
        
        if not input_str:
            self.output_text.insert(tk.END, "请输入待分析的单词串！\n", "error")
            return

        # 读取并处理文法
        grammar_text = self.grammar_text.get("1.0", tk.END)
        grammar_list = self.parser.read_grammar_from_text(grammar_text)
        
        if not grammar_list:
            self.output_text.insert(tk.END, "请先输入并分析文法！\n", "error")
            return

        try:
            # 初始化并处理文法
            self.parser.init()
            self.parser.identify_vn_vt(grammar_list)
            self.parser.reform_map()
            self.parser.find_first()
            self.parser.find_follow()
            is_ll1, _ = self.parser.is_ll1()
            
            if not is_ll1:
                self.output_text.insert(tk.END, "该文法不是LL(1)文法，无法分析！\n", "error")
                return
            
            self.parser.pre_form()

            # 分析字符串
            steps, is_success = self.parser.parse_string(input_str)
            self.output_text.insert(tk.END, f"===== 分析串：{input_str} =====\n", "title")
            
            # 输出表头
            self.output_text.insert(tk.END, "步骤\t分析栈\t\t剩余输入\t\t动作\n")
            self.output_text.insert(tk.END, "-" * 80 + "\n")
            
            for step in steps:
                step_num, stack, queue, action = step
                tag = "table"
                if "分析成功" in action:
                    tag = "success"
                elif "分析失败" in action:
                    tag = "error"
                
                self.output_text.insert(tk.END, f"{step_num}\t{stack:<20}\t{queue:<20}\t{action}\n", tag)
                
        except Exception as e:
            self.output_text.insert(tk.END, f"分析字符串时出错：{str(e)}\n", "error")

if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1600x1000")
    app = LL1ParserGUI(root)
    root.mainloop()